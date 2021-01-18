import os

import openpyxl

from misc.constants import GAUSS_SEIDEL


class ExcelWriter:
    def __init__(self, variable_dict: dict = None, file_name="output_in_tabular_format.xls"):
        self.file_name = file_name
        self.__clear_workbook()
        self.variable_names = ['' for i in range(len(variable_dict))]
        for key in variable_dict.keys():
            self.variable_names[variable_dict[key]] = key

    def write(self, matrix_solver):
        if matrix_solver.name == GAUSS_SEIDEL:
            self.__write_multi_iteration_algorithm(matrix_solver)
        else:
            self.__write_one_iteration_algorithm(matrix_solver)

    def __write_multi_iteration_algorithm(self, matrix_solver):
        ws = self.wb.create_sheet(matrix_solver.name)

        if matrix_solver.error != "":
            ws.cell(row=1, column=1, value=matrix_solver.error)
        else:
            ws.cell(row=1, column=1, value="Iteration")
            for i in range(len(self.variable_names)):
                ws.cell(row=1, column=i + 2, value=self.variable_names[i])
            for i in range(len(self.variable_names)):
                ws.cell(row=1, column=len(self.variable_names) + i + 2, value=f'Err({self.variable_names[i]})')
            ws.cell(row=1, column=2 * len(self.variable_names) + 2, value=f'Max Err')

            for i in range(len(matrix_solver.iterations_list)):
                ws.cell(row=i + 2, column=1, value=i + 1)
                for j in range(len(self.variable_names)):
                    ws.cell(row=i + 2, column=j + 2, value=matrix_solver.iterations_list[i].values[j])
                    ws.cell(row=i + 2, column=len(self.variable_names) + j + 2,
                            value=matrix_solver.iterations_list[i].errors[j])
                ws.cell(row=i + 2, column=2 * len(self.variable_names) + 2,
                        value=matrix_solver.iterations_list[i].max_error)
        self.wb.save(self.file_name)

    def __write_one_iteration_algorithm(self, matrix_solver):
        ws = self.wb.create_sheet(matrix_solver.name)
        if matrix_solver.error != "":
            ws.cell(row=1, column=1, value=matrix_solver.error)
        else:
            for i in range(len(self.variable_names)):
                ws.cell(row=1, column=i + 1, value=self.variable_names[i])
                ws.cell(row=2, column=i + 1, value=matrix_solver.result[i])
        self.wb.save(self.file_name)

    def __clear_workbook(self):
        if os.path.exists(self.file_name):
            os.remove(self.file_name)
        self.wb = openpyxl.Workbook()
        self.wb.save(self.file_name)

    def close(self):
        self.wb.remove_sheet(self.wb['Sheet'])
        self.wb.save(self.file_name)
        self.wb.close()
